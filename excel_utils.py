import json
import os
import time
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Union

def create_excel(data: dict, file_name: str) -> str:
    """Create a professional Excel file from extracted data, including a hidden raw data cell."""
    wb = Workbook()

    # --- Helper function for styling ---
    def style_sheet(ws):
        for col in ws.columns:
            max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_length + 2)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Sheet 1: Report Details ---
    ws_report = wb.active
    ws_report.title = "Report Details"
    ws_report.append(["Field", "Value"])
    report_fields = [
        ("Report ID", data.get("report_id")),
        ("Classification", data.get("report_classification")),
        ("Location", data.get("report_location")),
        ("Time", data.get("report_time")),
        ("Police Agency", data.get("police_agency")),
        ("Report Filename", data.get("report_filename")),
        ("Process Time (sec)", data.get("process_time"))
    ]
    for field, value in report_fields:
        ws_report.append([field, value])
    style_sheet(ws_report)

    # --- Sheet 2: Persons ---
    ws_persons = wb.create_sheet(title="Persons")
    ws_persons.append([
        "Name", "Role", "Date of Birth", "Age", "Address", "Race",
        "First Seen Bate", "Bates References", "Narrative Snippet", "Relationships"
    ])
    for p in data.get("persons", []):
        relationships = "; ".join([f"{r['name']} ({r['relationship']})" for r in p.get("relationships", [])])
        ws_persons.append([
            p.get("name"),
            p.get("role"),
            p.get("date_of_birth"),
            p.get("age"),
            p.get("address"),
            p.get("race"),
            p.get("first_seen_bate"),
            ", ".join(p.get("bates_references", [])),
            p.get("narrative_snippet"),
            relationships
        ])
    style_sheet(ws_persons)

    # --- Sheet 3: Officers ---
    ws_officers = wb.create_sheet(title="Officers")
    ws_officers.append([
        "Name", "Rank", "Badge Number", "First Seen Bate", "Bates References", "Narrative Snippet"
    ])
    for o in data.get("officers", []):
        ws_officers.append([
            o.get("name"),
            o.get("rank"),
            o.get("badge_number"),
            o.get("first_seen_bate"),
            ", ".join(o.get("bates_references", [])),
            o.get("narrative_snippet")
        ])
    style_sheet(ws_officers)

    # --- Sheet 4: Officials ---
    ws_officials = wb.create_sheet(title="Officials")
    ws_officials.append([
        "Name", "Title", "Agency/Court", "First Seen Bate", "Bates References", "Narrative Snippet"
    ])
    for off in data.get("officials", []):
        ws_officials.append([
            off.get("name"),
            off.get("title"),
            off.get("agency_or_court"),
            off.get("first_seen_bate"),
            ", ".join(off.get("bates_references", [])),
            off.get("narrative_snippet")
        ])

    # --- Wrap narrative cells ---
    for row in ws_persons.iter_rows(min_row=2):
        row[8].alignment = Alignment(wrap_text=True)
    for row in ws_officers.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True)
    for row in ws_officials.iter_rows(min_row=2):
        row[5].alignment = Alignment(wrap_text=True)

    style_sheet(ws_officials)

    # --- Save workbook ---
    excel_path = f"{os.path.splitext(file_name)[0]}_{int(time.time())}.xlsx"
    wb.save(excel_path)
    return excel_path

def rebuild_data_from_excel(excel_source: Union[str, BytesIO]) -> dict:
    """Reconstruct JSON from the actual Excel content."""
    wb = load_workbook(excel_source, data_only=True)
    data = {}

    # --- Sheet: Persons ---
    ws_persons = wb["Persons"]
    persons = []
    for row in ws_persons.iter_rows(min_row=2, values_only=True):
        name, role, dob, age, address, race, first_bate, bates_refs, narrative, relationships_str = row
        relationships = []
        if relationships_str:
            for rel in relationships_str.split(";"):
                if "(" in rel and ")" in rel:
                    r_name, r_type = rel.strip().rsplit("(", 1)
                    relationships.append({"name": r_name.strip(), "relationship": r_type.rstrip(")")})
        persons.append({
            "name": name,
            "role": role,
            "date_of_birth": dob,
            "age": age,
            "address": address,
            "race": race,
            "first_seen_bate": first_bate,
            "bates_references": [b.strip() for b in bates_refs.split(",")] if bates_refs else [],
            "narrative_snippet": narrative,
            "relationships": relationships
        })
    data["persons"] = persons

    # --- Sheet: Officers ---
    ws_officers = wb["Officers"]
    officers = []
    for row in ws_officers.iter_rows(min_row=2, values_only=True):
        name, rank, badge, first_bate, bates_refs, narrative = row
        officers.append({
            "name": name,
            "rank": rank,
            "badge_number": badge,
            "first_seen_bate": first_bate,
            "bates_references": [b.strip() for b in bates_refs.split(",")] if bates_refs else [],
            "narrative_snippet": narrative
        })
    data["officers"] = officers

    # --- Sheet: Officials ---
    ws_officials = wb["Officials"]
    officials = []
    for row in ws_officials.iter_rows(min_row=2, values_only=True):
        name, title, agency, first_bate, bates_refs, narrative = row
        officials.append({
            "name": name,
            "title": title,
            "agency_or_court": agency,
            "first_seen_bate": first_bate,
            "bates_references": [b.strip() for b in bates_refs.split(",")] if bates_refs else [],
            "narrative_snippet": narrative
        })
    data["officials"] = officials

    # --- Sheet: Report Details ---
    ws_report = wb["Report Details"]
    data["report_id"] = ws_report["B2"].value
    data["report_classification"] = ws_report["B3"].value
    data["report_location"] = ws_report["B4"].value
    data["report_time"] = ws_report["B5"].value
    data["police_agency"] = ws_report["B6"].value
    data["report_filename"] = ws_report["B7"].value
    data["process_time"] = ws_report["B8"].value

    return data

def create_excel_combined_parties(data: dict, file_name: str) -> str:
    """Create Excel file with:
       1) All Parties (combined structure with color-coded rows)
       2) Report Details (second worksheet)
    """

    wb = Workbook()

    # Define color fills for each party type
    person_fill = PatternFill(start_color="EEF4DE", end_color="EEF4DE", fill_type="solid")
    officer_fill = PatternFill(start_color="EBDDE2", end_color="EBDDE2", fill_type="solid")
    official_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # ==========================================================
    # SHEET 1: ALL PARTIES
    # ==========================================================
    ws = wb.active
    ws.title = "All Parties"

    headers = [
        "Party Type",
        "Name",
        "Role",
        "Date of Birth",
        "Age",
        "Address",
        "First Seen Bate",
        "Bates References",
        "Narrative Snippet",
        "Race",
        "Agency / Court",
        "Badge Number",
        "Relationships"
    ]

    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---------------- PERSONS ----------------
    for p in data.get("persons", []):
        relationships = "; ".join(
            [f"{r['name']} ({r['relationship']})" for r in p.get("relationships", [])]
        )

        ws.append([
            "Person",
            p.get("name"),
            p.get("role"),
            p.get("date_of_birth"),
            p.get("age"),
            p.get("address"),
            p.get("first_seen_bate"),
            ", ".join(p.get("bates_references", [])),
            p.get("narrative_snippet"),
            p.get("race"),
            None,
            None,
            relationships
        ])
        
        # Apply person color to the entire row
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill = person_fill

    ws.append([])

    # ---------------- OFFICERS ----------------
    for o in data.get("officers", []):
        ws.append([
            "Officer",
            o.get("name"),
            o.get("rank"),
            None,
            None,
            None,
            o.get("first_seen_bate"),
            ", ".join(o.get("bates_references", [])),
            o.get("narrative_snippet"),
            None,
            None,
            o.get("badge_number"),
            None
        ])
        
        # Apply officer color to the entire row
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill = officer_fill

    ws.append([])

    # ---------------- OFFICIALS ----------------
    for off in data.get("officials", []):
        ws.append([
            "Official",
            off.get("name"),
            off.get("title"),
            None,
            None,
            None,
            off.get("first_seen_bate"),
            ", ".join(off.get("bates_references", [])),
            off.get("narrative_snippet"),
            None,
            off.get("agency_or_court"),
            None,
            None
        ])
        
        # Apply official color to the entire row
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill = official_fill

    # Wrap narrative column
    for row in ws.iter_rows(min_row=2):
        row[8].alignment = Alignment(wrap_text=True)

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_length + 2)

    # ==========================================================
    # SHEET 2: REPORT DETAILS
    # ==========================================================
    ws_report = wb.create_sheet(title="Report Details")
    ws_report.append(["Field", "Value"])

    report_fields = [
        ("Report ID", data.get("report_id")),
        ("Classification", data.get("report_classification")),
        ("Location", data.get("report_location")),
        ("Time", data.get("report_time")),
        ("Police Agency", data.get("police_agency")),
        ("Report Filename", data.get("report_filename")),
        ("Process Time (sec)", data.get("process_time")),
    ]

    for field, value in report_fields:
        ws_report.append([field, value])

    for cell in ws_report[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws_report.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_report.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_length + 2)

    # ==========================================================
    # SAVE
    # ==========================================================
    excel_path = f"{os.path.splitext(file_name)[0]}.xlsx"
    wb.save(excel_path)

    return excel_path

def rebuild_data_from_combined_excel(excel_source: Union[str, BytesIO]) -> dict:
    """Reconstruct structured JSON from:
       - All Parties sheet
       - Report Details sheet
    """

    wb = load_workbook(excel_source, data_only=True)

    data = {
        "persons": [],
        "officers": [],
        "officials": []
    }

    # ==========================================================
    # READ ALL PARTIES
    # ==========================================================
    ws = wb["All Parties"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        (
            party_type,
            name,
            role,
            dob,
            age,
            address,
            first_bate,
            bates_refs,
            narrative,
            race,
            agency,
            badge,
            relationships_str
        ) = row

        bates_list = [b.strip() for b in bates_refs.split(",")] if bates_refs else []

        if party_type == "Person":
            relationships = []

            if relationships_str:
                for rel in relationships_str.split(";"):
                    if "(" in rel and ")" in rel:
                        r_name, r_type = rel.strip().rsplit("(", 1)
                        relationships.append({
                            "name": r_name.strip(),
                            "relationship": r_type.rstrip(")")
                        })

            data["persons"].append({
                "name": name,
                "role": role,
                "date_of_birth": dob,
                "age": age,
                "address": address,
                "race": race,
                "first_seen_bate": first_bate,
                "bates_references": bates_list,
                "narrative_snippet": narrative,
                "relationships": relationships
            })

        elif party_type == "Officer":
            data["officers"].append({
                "name": name,
                "rank": role,
                "badge_number": badge,
                "first_seen_bate": first_bate,
                "bates_references": bates_list,
                "narrative_snippet": narrative
            })

        elif party_type == "Official":
            data["officials"].append({
                "name": name,
                "title": role,
                "agency_or_court": agency,
                "first_seen_bate": first_bate,
                "bates_references": bates_list,
                "narrative_snippet": narrative
            })

    # ==========================================================
    # READ REPORT DETAILS
    # ==========================================================
    ws_report = wb["Report Details"]

    data["report_id"] = ws_report["B2"].value
    data["report_classification"] = ws_report["B3"].value
    data["report_location"] = ws_report["B4"].value
    data["report_time"] = ws_report["B5"].value
    data["police_agency"] = ws_report["B6"].value
    data["report_filename"] = ws_report["B7"].value
    data["process_time"] = ws_report["B8"].value

    return data